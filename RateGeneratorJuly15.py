#!/usr/bin/env python
# coding: utf-8

# In[1]:


# pip install thefuzz pandas


# In[2]:


# pip install pyarrow


# In[3]:


# pip install pandas-gbq --upgrade


# In[4]:


# In[5]:


# pip install openpyxl


# In[6]:


import os
import pandas as pd
from rapidfuzz import process

from pathlib import Path

FOLDER_PATH = Path().resolve() / "RateSheet_Project" / "RateSheetFiles"
files = list(FOLDER_PATH.glob("*.xlsx"))

# Read all Excel files in the specified folder
files = [os.path.join(FOLDER_PATH, f) for f in os.listdir(FOLDER_PATH) if f.endswith(".xlsx")]
dfs = {path: pd.read_excel(path, dtype=str) for path in files}

print(f"✅ found {len(files)} Excel files: {files}")


# In[7]:


# ✅ Read the first 10 rows of each file to detect the header row
def detect_header_row(df):
    """ Check 'POL' for header row """
    for i in range(len(df)):
        row_str = " ".join(df.iloc[i].astype(str))
        print(f"🔍 Check {i} row: {row_str}") 
        if "POL" in row_str.upper():  # capitalize for case-insensitive match
            print(f"✅ FOUND 'POL' at {i} row")
            return i
    print("⚠️ Cannot find 'POL'returning default header row 0")
    return 0 

# read the first 10 rows of each file to detect the header row
header_rows = {str(path): detect_header_row(pd.read_excel(path, nrows=10, header=None)) for path in files}

# ✅ based on detected header rows, read the full files
dfs = {path: pd.read_excel(path, header=header_rows[path]) for path in files}

# ✅ Merge duplicate columns
def merge_duplicate_columns(df):
    if 'remark' in df.columns and 'remark' in df.columns:
        df['remark'] = df['remark'].fillna(df['remark'])  
        df = df.drop(columns=['remark']) 
    return df

# ✅ Apply the merge function to all dataframes
print("\n📌Head row found, files with head row number：")
for path, header_row in header_rows.items():
    print(f"📄 {os.path.basename(path)} - header row: {header_row}")


# In[ ]:


import pandas as pd

# unify the column name for "remarks" related columns across all dataframes
def unify_remark_column_name(df):
    # possible column names for remarks
    possible_columns = ["Rate remarks", "rate remarks", "REMARK", "REMARKS", "Remark", "Remarks", "remark", "remarks", "Rate_remarks", "RATE_REMARKS", "RATE_REMARK", "Rate_remark", "rate_remarks"]

    # validate and rename the column if it exists
    for col in possible_columns:
        if col in df.columns:  #if the column exists in the dataframe
            df = df.rename(columns={col: 'remark'})
            break  # exit loop after renaming
    return df

# dictionary to store processed dataframes
for file, df in dfs.items():
    df = unify_remark_column_name(df)  

    print(f"📄 {file} Cleaned Head Row Name:", df.columns.tolist())


# In[ ]:


import re
# Header cleaning function
def clean_column_names(df):
    new_columns = []
    for col in df.columns:
        col = col.strip()  # strip whitespace
        col = col.replace(" ", "_")  
        col = col.replace("/", "_")  
        col = col.replace("(", "")   
        col = col.replace(")", "")   
        col = re.sub(r'_+', '_', col)  
        col = col.replace("'", "")

        # avoid leading numbers
        col = re.sub(r"^(\d+)([A-Z]+)(\.\d+)?$", r"\2\1\3", col)  # 例：20GP.1 → GP20.1

        new_columns.append(col)

    df.columns = new_columns
    return df

# unify the column names and clean them

for file, df in dfs.items():
    df = clean_column_names(df)
    
    df = unify_remark_column_name(df)

    df = df.rename(columns={
        'CARRIER': 'Carrier',
        'DESTINATION': 'Destination',
        'T_T': 'T_T_TO_POD',
        'EFFECTIVE_DATE': 'Effective_Date',
        'EXPIRY_DATE': 'Expiring_Date'
    })
    
    df = df.loc[:, ~df.columns.duplicated()]
    dfs[file] = df
    print(f"📄 {file} Cleaned Head Row Name:", df.columns.tolist())


# In[ ]:


from rapidfuzz import process

# ports names "Yantian, Shenzhen": ["yantian", "YTN"],
port_aliases = {
    "SHENZHEN, GUANGDONG": ["shekou", "SHK", "yantian", "YTN"],
    "JIUJIANG, GUANGDONG": ["jiujiang", "JJG"],
    "HONG KONG": ["hong kong", "HKG"],
    "ZHUHAI, GUANGDONG": ["zhuhai", "ZUH"],
    "ZHONGSHAN, GUANGDONG": ["zhongshan", "ZSN"],
    "NANSHA, GUANGDONG": ["nansha", "NSA"],
    "HUANGPU, GUANGDONG": ["huangpu", "HUP"],
    "XIAMEN, FUJIAN": ["xiamen", "XMN"],
    "FUZHOU, FUJIAN": ["fuzhou", "FOC"],
    "ZHENJIANG, JIANGSU": ["zhenjiang", "ZJG"],
    "ZHAPU, ZHEJIANG": ["zhapu", "ZPU"],
    "ZHANGJIAGANG, JIANGSU": ["zhangjiagang", "ZJG"],
    "YUEYANG, HUNAN": ["yueyang", "YYG"],
    "YICHANG, HUBEI": ["yichang", "YIC"],
    "YANGZHOU, JIANGSU": ["yangzhou", "YZH"],
    "WUHU, ANHUI": ["wuhu", "WHU"],
    "WUHAN, HUBEI": ["wuhan", "WUH"],
    "SHANGHAI": ["shanghai", "SHA"],
    "NINGBO, ZHEJIANG": ["ningbo", "NGB"],
    "NANTONG, JIANGSU": ["nantong", "NTG"],
    "NANJING, JIANGSU": ["nanjing", "NKG"],
    "NANCHANG, JIANGXI": ["nanchang", "NCG"],
    "CHANGZHOU, JIANGSU": ["changzhou", "CZH"],
    "CHANGSHA, HUNAN": ["changsha", "CSX"],
    "ANQING, ANHUI": ["anqing", "AQG"],
    "XINGANG, TIANJIN": ["xingang", "XGG"],
    "QINGDAO, SHANDONG": ["qingdao", "QDG"],
    "DALIAN, LIAONING": ["dalian", "DLC"],
    "YOKOHAMA, JAPAN": ["yokohama", "YOK"],
    "VUNG TAU, VIETNAM": ["vung tau", "VUT"],
    "VISAKHAPATNAM, INDIA": ["visakhapatnam", "VSK"],
    "TUTICORIN, INDIA": ["tuticorin", "TUT"],
    "TOKYO, JAPAN": ["tokyo", "TYO"],
    "TAOYUAN, TAIWAN": ["taoyuan", "TYN"],
    "TANJUNG PELEPAS, MALAYSIA": ["tanjung pelepas", "TPP"],
    "TAIPEI, TAIWAN": ["taipei", "TPE"],
    "TAICHUNG, TAIWAN": ["taichung", "TXG"],
    "SURABAYA, INDONESIA": ["surabaya", "SUB"],
    "SUBIC BAY, PHILIPPINES": ["subic bay", "SUB"],
    "SINGAPORE": ["singapore", "SIN"],
    "SIHANOUKVILLE, CAMBODIA": ["sihanoukville", "SIH"],
    "SHIMIZU, JAPAN": ["shimizu", "SZU"],
    "SEMARANG, INDONESIA": ["semarang", "SRG"],
    "QUI NHON, VIETNAM": ["qui nhon", "QNH"],
    "PORT KLANG, MALAYSIA": ["port klang", "PKL"],
    "PHNOM PENH, CAMBODIA": ["phnom penh", "PNH"],
    "PENANG, MALAYSIA": ["penang", "PEN"],
    "PASIR GUDANG, MALAYSIA": ["pasir gudang", "PGU"],
    "PALEMBANG, INDONESIA": ["palembang", "PLM"],
    "OSAKA, JAPAN": ["osaka", "OSA"],
    "NHAVA SHEVA, INDIA": ["nhava sheva", "NSH"],
    "NAGOYA, JAPAN": ["nagoya", "NGO"],
    "MUNDRA, INDIA": ["mundra", "MUN"],
    "MOJI, JAPAN": ["moji", "MOJ"],
    "MANILA, PHILIPPINES": ["manila", "MNL"],
    "MANILA NORTH HARBOUR": ["manila north harbour", "MNH"],
    "LAT KRABANG, THAILAND": ["lat krabang", "LKB"],
    "LAEM CHABANG, THAILAND": ["laem chabang", "LCH"],
    "KOLKATA(EX CALCUTTA), INDIA": ["kolkata(ex calcutta)", "CCU"],
    "KOBE, JAPAN": ["kobe", "UKB"],
    "KEELUNG, TAIWAN": ["keelung", "KEL"],
    "KARACHI, PAKISTAN": ["karachi", "KHI"],
    "KAOHSIUNG, TAIWAN": ["kaohsiung", "KHH"],
    "JAKARTA, INDONESIA": ["jakarta", "JKT"],
    "HOCHIMINH CITY, VIETNAM": ["hochiminh city", "SGN"],
    "HAKATA, JAPAN": ["hakata", "HAK"],
    "HAIPHONG, VIETNAM": ["haiphong", "HPH"],
    "DAVAO, PHILIPPINES": ["davao", "DVO"],
    "DANANG, VIETNAM": ["danang", "DAD"],
    "COLOMBO, SRI LANKA": ["colombo", "CMB"],
    "COCHIN, INDIA": ["cochin", "COK"],
    "CHATTOGRAM, BANGLADESH": ["chattogram", "CGP"],
    "CHENNAI, INDIA": ["chennai", "MAA"],
    "CEBU, PHILIPPINES": ["cebu", "CEB"],
    "CAI MEP, VIETNAM": ["cai mep", "CMV"],
    "BUSAN, KOREA": ["busan", "PUS"],
    "BELAWAN, INDONESIA": ["belawan", "BLW"],
    "BATAM, INDONESIA": ["batam", "BTH"],
    "BANGKOK, THAILAND": ["bangkok", "BKK"],
    "PANJANG, INDONESIA": ["panjang", "PNJ"],
    "PIPAVAV (VICTOR) PORT, INDIA": ["pipavav (victor) port", "PIP"],
    "HAZIRA, INDIA": ["hazira", "HZR"],
    "KATTUPALLI, INDIA": ["kattupalli", "KTP"]
}


# In[ ]:


from rapidfuzz import process

def fuzzy_match_pol(pol):
    if pd.isna(pol) or pol.strip() == "":
        return pol
    
    pol = pol.upper().strip() 
    port_list = list(port_aliases.keys())  
    
    for full_name, aliases in port_aliases.items():
        if any(code in pol for code in aliases):
            return full_name 

    # `fuzzy match`
    match = process.extractOne(pol, port_list)

    if match:
        best_match, score, _ = match 
        return best_match if score > 75 else pol
    else:
        return pol



# In[ ]:


# turn the POL column in each dataframe to its full name
for path, df in dfs.items():
    if "POL" in df.columns:
        print(f"\n📂 processing file: {path}")
        print("🔍 Old POL:", df["POL"].dropna().unique()[:10]) 
        df["POL"] = df["POL"].apply(fuzzy_match_pol)
        print("✅ Cleaned POL:", df["POL"].dropna().unique()[:10])
    else:
        print(f"⚠️ file has no POL columns: {path}")


# In[ ]:


import pandas as pd
import os
import glob
import re
from datetime import datetime
from pathlib import Path

def standardize_date_columns(df, filename):
    year_match = re.search(r'202[0-9]{1}', filename)
    expected_year = int(year_match.group()) if year_match else 2025
    for col in df.columns:
        if "date" in col.lower():
            print(f"Column {col}, Type: {df[col].dtype}")
            print(f"Old3: {df[col].head(3).tolist()}")
            if pd.api.types.is_numeric_dtype(df[col]):
                df[col] = pd.to_datetime("1899-12-30") + pd.to_timedelta(df[col], unit="D")
                df[col] = df[col].apply(lambda x: x.replace(year=expected_year) if pd.notna(x) and (x.year < expected_year - 1 or x.year > expected_year + 1) else x)
            else:
                df[col] = pd.to_datetime(df[col], errors="coerce", format=None)
                def parse_manual(date_str):
                    if pd.isna(date_str) or str(date_str).strip() in ["", "NIL", "-", "—"]:
                        return pd.NaT
                    for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%b-%Y", "%Y.%m.%d"]:
                        try:
                            parsed = datetime.strptime(str(date_str), fmt)
                            if parsed.year < expected_year - 1 or parsed.year > expected_year + 1:
                                return parsed.replace(year=expected_year)
                            return parsed
                        except ValueError:
                            continue
                    return pd.NaT
                df[col] = df[col].apply(lambda x: parse_manual(x) if pd.isna(x) else x)
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%Y-%m-%d")
            print(f"Cleaned3: {df[col].head(3).tolist()}")
    return df

def extract_formula_values(filepath):
    from openpyxl import load_workbook
    import pandas as pd
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    header = data[0]
    df = pd.DataFrame(data[1:], columns=header)
    return df

# dictionary to store processed dataframes
folder_path = Path().resolve() / "RateSheet_Project" / "RateSheetFiles" / "Cleaned"

for file in dfs:
    print(f"\n⏳ Cleaned Date: {file}")
    dfs[file] = standardize_date_columns(dfs[file], file)


# In[ ]:


from rapidfuzz import process

# carriers names and their possible aliases or SCAC codes
carrier_aliases = {
    "WANHAI": ["WHLC", "WANHAI", "WHAI"],
    "TSL": ["TSYN", "TSL"],
    "SMLM": ["SML", "SMLM"],
    "YML": ["YMJA", "YML"],
    "MSC": ["MEDU", "MSC"],
    "OOCL": ["OOLU", "OOCL"],
    "ONE": ["ONEY", "ONE"],
    "EMC": ["EGLV", "EMC"],
    "COSCO": ["COSU", "COSCO"],
    "HMM": ["HDMU", "HMM"],
    "HPL": ["HLCU", "HPL"],
    "CMA": ["CMDU", "CMA", "CMU"],
    "ZIM": ["ZIMU", "ZIM"],
    "SLS": ["SLS"],
    "HEDE": ["HEDE"],
    "MATS": ["MATS", "MATSON"],
}


# In[ ]:


def fuzzy_match_carrier(val):
    if pd.isna(val) or str(val).strip() == "":
        return val

    val = str(val).strip().upper()
    standard_names = list(carrier_aliases.keys())

    # match against standard names directly
    for standard, aliases in carrier_aliases.items():
        if val in aliases:
            return standard
    
    # fuzzy match
    match = process.extractOne(val, standard_names)
    if match:
        best_match, score, _ = match
        return best_match if score > 75 else val
    else:
        return val

for path, df in dfs.items():
    if "Carrier" in df.columns:
        print(f"\n📄 file: {path}")
        print("🔍 old:", df["Carrier"].dropna().unique()[:10])
        df["Carrier"] = df["Carrier"].apply(fuzzy_match_carrier)
        print("✅ cleaned:", df["Carrier"].dropna().unique()[:10])


# In[ ]:


from collections import defaultdict
import re

# Create a mapping of keywords to standardized city names
city_mapping_keywords = {
    "LAX/LGB": "LOS ANGELES, LAX, LGB, LAX/LGB, LONG BEACH",
    "CHICAGO, IL": "CHICAGO, JOLIET, CHI, USCHI",
    "NEW YORK, NY": "NEW YORK, NYC, USNYC",
    "DALLAS, TX": "DALLAS, USDAL, DAL",
    "HOUSTON, TX": "HOUSTON",
    "SEATTLE, WA": "SEATTLE",
    "TACOMA, WA":"TACOMA",
    "OAKLAND, CA": "OAKLAND",
    "MIAMI, FL": "MIAMI",
    "HONOLULU, HI": "HONOLULU",
    "CLEVELAND, OH": "CLEVELAND",
    "BALTIMORE, MD": "BALTIMORE",
    "CHARLESTON, SC": "CHARLESTON",
    "PORTLAND, OR": "PORTLAND",
    "MEMPHIS, TN": "MEMPHIS",
    "SAVANNAH, GA": "SAVANNAH",
    "PHILADELPHIA, PA": "PHILADELPHIA",
    "ATLANTA, GA": "ATLANTA",
    "INDIANAPOLIS, IN": "INDIANAPOLIS",
    "DETROIT, MI": "DETROIT",
    "TAMPA, FL": "TAMPA",
    "SAINT LOUIS, MO": "SAINT LOUIS",
    "JACKSONVILLE, FL": "JACKSONVILLE",
    "KANSAS CITY, MO": "KANSAS CITY",
    "MINNEAPOLIS, MN": "MINNEAPOLIS",
    "CINCINNATI, OH": "CINCINNATI",
    "DENVER, CO": "DENVER",
    "PHOENIX, AZ": "PHOENIX",
    "SALT LAKE CITY, UT": "SALT LAKE CITY",
    "NASHVILLE, TN": "NASHVILLE",
    "OMAHA, NE": "OMAHA",
    "PITTSBURGH, PA": "PITTSBURGH",
    "BOSTON, MA": "BOSTON",
    "BUFFALO, NY": "BUFFALO",
    "LOUISVILLE, KY": "LOUISVILLE",
    "EL PASO, TX": "EL PASO",
    "COLUMBUS, OH": "COLUMBUS",
    "HILO, HI": "HILO",
    "KAHULUI, HI": "KAHULUI",
    "SASKATOON, CANADA": "SASKATOON",
    "CALGARY, CANADA": "CALGARY",
    "EDMONTON, CANADA": "EDMONTON",
    "VANCOUVER, CANADA": "VANCOUVER",
    "TORONTO, CANADA": "TORONTO",
    "MONTREAL, CANADA": "MONTREAL",
    "PRINCE RUPERT, CANADA": "PRINCE RUPERT",
    "HALIFAX, CANADA": "HALIFAX",
    "REGINA, CANADA": "REGINA",
    "WINNIPEG, CANADA": "WINNIPEG",
}


# In[ ]:


def fuzzy_match_city(val):
    if pd.isna(val) or str(val).strip() == "":
        return val

    val = str(val).strip().upper()
    val_city = re.split(r"[,-]", val)[0].strip()  
    standard_names = list(city_mapping_keywords.keys())

    # match against standard names directly
    for standard, aliases in city_mapping_keywords.items():
        if val_city in aliases:
            return standard

    # fuzzy match
    all_aliases = [alias for aliases in city_mapping_keywords.values() for alias in aliases]
    match = process.extractOne(val_city, all_aliases)
    if match:
        best_match, score, _ = match
        for standard, aliases in city_mapping_keywords.items():
            if best_match in aliases:
                return standard if score > 75 else val
    return val

# apply to all dataframes
for path, df in dfs.items():
    if "Destination" in df.columns:
        print(f"\n📄 file: {path}")
        print("🔍 old:", df["Destination"].dropna().unique()[:10])
        df["Destination"] = df["Destination"].apply(fuzzy_match_city)
        print("✅ cleaned:", df["Destination"].dropna().unique()[:10])


# In[ ]:


from google.cloud import bigquery
from google.oauth2 import service_account

# ✅ Set up Google Cloud credentials
BASE_DIR = Path().resolve()
json_files = list(BASE_DIR.rglob("*.json"))
if not json_files:
    raise FileNotFoundError("❌ not found .json credentials file")
credentials_path = json_files[0]
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = str(credentials_path)

print("GOOGLE_APPLICATION_CREDENTIALS =", os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"))


# In[ ]:


# ✅ BigQuery client setup
project_id = "rate-sheet-sql-465312"
dataset_id = "ratesheet_processing_dataset"
credentials = service_account.Credentials.from_service_account_file(credentials_path)
client = bigquery.Client(credentials=credentials, project=project_id)

for dataset in client.list_datasets():
    print("✅ found dataset：", dataset.dataset_id)


# In[27]:


BASE_DIR = Path().resolve()
FOLDER_PATH = BASE_DIR / "RateSheet_Project" / "RateSheetFiles" / "Cleaned"
excel_files = list(FOLDER_PATH.glob("*.xlsx"))

def clean_table_name(file_path):
    print(f"📂 file accurate path: {file_path}")
    name = Path(file_path).stem.lower()
    name = re.sub(r"[^a-z0-9_]", "_", name)
    name = re.sub(r"_+", "_", name)
    name = name.strip("_")
    print("📥 file name：", name)
    return name

for file in excel_files:
    table_name = clean_table_name(file)
    print("✅ outside file name：", table_name)


# In[28]:


# ✅ found old Excel files
FOLDER_PATH = BASE_DIR / "RateSheet_Project" / "RateSheetFiles"
excel_files = list(FOLDER_PATH.glob("*.xlsx"))
if not excel_files:
    raise FileNotFoundError("❌ no Excel found")
else:
    # ✅ delete old tables
    tables = client.list_tables(dataset_id)
    for table in tables:
        table_ref = f"{project_id}.{dataset_id}.{table.table_id}"
        client.delete_table(table_ref, not_found_ok=True)
        print(f"🗑 Sheet deleted：{table_ref}")


# In[ ]:


import os

output_folder = Path().resolve() / "RateSheet_Project" / "RateSheetFiles" / "Cleaned"
os.makedirs(output_folder, exist_ok=True)
print(f"✅ output: {output_folder}")

for path, df in dfs.items():
    filename = os.path.basename(path)
    output_path = output_folder / f"cleaned_{filename}"
    df.to_excel(output_path, index=False)
    print(f"✅ save: {output_path}")


# In[ ]:


from datetime import datetime, timedelta
from google.cloud import bigquery
import pytz

from pathlib import Path
excel_files = list((Path().resolve() / "RateSheet_Project" / "RateSheetFiles" / "Cleaned").glob("*.xlsx"))

for file in excel_files:
    table_name = clean_table_name(file)
    table_id = f"{project_id}.{dataset_id}.{table_name}"
    print(f"⏳ read：{file}")

    df = pd.read_excel(file)

    for col in df.columns:
        if "date" not in col.lower():
            df[col] = df[col].apply(
                lambda x: pd.NA if str(x).strip().upper() in ["", "NIL", "-", "—"] else x
            )

    keep_columns = [
        'POL', 'Carrier', 'T_T_TO_POD', 'Destination',
        'Effective_Date', 'Expiring_Date',
        'GP20', 'GP40', 'HQ40', 'HQ45',
        'COMM', 'COMM_DETAILS',
        'COMMODITY', 'remark'
    ]

    df = df[[col for col in keep_columns if col in df.columns]]

    job_config = bigquery.LoadJobConfig()
    job_config.write_disposition = bigquery.WriteDisposition.WRITE_TRUNCATE

    job = client.load_table_from_dataframe(df, table_id, job_config=job_config)
    job.result()

    print(f"✅ Successfully uploaded → {table_id}")


# In[ ]:





# In[ ]:




